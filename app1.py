# import openpyxl and tkinter modules
from tkinter.ttk import Style

import tk as tk
from Tools.demo.spreadsheet import center
from openpyxl import *
from tkinter import *


# globally declare wb and sheet variable

# opening the existing excel file
wb = load_workbook(r'F:\COLLAGE FILES\2nd YEAR\SEM 2\Hackathon\students.xlsx')

# create the sheet object
sheet = wb.active

def excel():
    # resize the width of columns in
    # excel spreadsheet
    sheet.column_dimensions['A'].width = 30
    sheet.column_dimensions['B'].width = 30
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 10
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 40
    sheet.column_dimensions['G'].width = 50
    sheet.column_dimensions['H'].width = 60
    sheet.column_dimensions['I'].width = 60

    # write given data to an excel spreadsheet
    # at particular location
    sheet.cell(row=1, column=1).value = "SurName"
    sheet.cell(row=1, column=2).value = "FirstName"
    sheet.cell(row=1, column=3).value = "LastName"
    sheet.cell(row=1, column=4).value = "Course"
    sheet.cell(row=1, column=5).value = "Contact Number"
    sheet.cell(row=1, column=6).value = "Email id"
    sheet.cell(row=1, column=7).value = "Address"
    sheet.cell(row=1, column=8).value = "strengths"
    sheet.cell(row=1, column=9).value = "weekness"

# Function to set focus (cursor)
def focus1(event1):
    # set focus on the course_field box
    fname_field.focus_set()


# Function to set focus
def focus2(event):
    # set focus on the sem_field box
    lname_field.focus_set()


# Function to set focus
def focus3(event):
    # set focus on the form_no_field box
    course_field.focus_set()


# Function to set focus
def focus4(event):
    # set focus on the contact_no_field box
    contact_no_field.focus_set()


# Function to set focus
def focus5(event):
    # set focus on the email_id_field box
    email_id_field.focus_set()


# Function to set focus
def focus6(event):
    # set focus on the address_field box
    address_field.focus_set()

def focus7(event):
    # set focus on the address_field box
    strengths_field.focus_set()

def focus8(event):
    # set focus on the address_field box
    weekness_field.focus_set()

# Function for clearing the
# contents of text entry boxes
def clear():
    # clear the content of text entry box
    sname_field.delete(0, END)
    fname_field.delete(0, END)
    lname_field.delete(0, END)
    course_field.delete(0, END)
    contact_no_field.delete(0, END)
    email_id_field.delete(0, END)
    address_field.delete(0, END)
    strengths_field.delete(0, END)
    weekness_field.delete(0, END)

# Function to take data from GUI
# window and write to an excel file
def insert():
    # if user not fill any entry
    # then print "empty input"
    if (sname_field.get() == "" and
            fname_field.get() == "" and
            lname_field.get() == "" and
            course_field.get() == "" and
            contact_no_field.get() == "" and
            email_id_field.get() == "" and
            address_field.get() == "" and
            strengths_field.get() == "" and
            weekness_field.get() == ""):

        print("empty input")

    else:

        # assigning the max row and max column
        # value upto which data is written
        # in an excel sheet to the variable
        current_row = sheet.max_row
        current_column = sheet.max_column


# get method returns current text
        # as string which we write into
        # excel spreadsheet at particular location
        sheet.cell(row=current_row + 1, column=1).value = sname_field.get()
        sheet.cell(row=current_row + 1, column=2).value = fname_field.get()
        sheet.cell(row=current_row + 1, column=3).value = lname_field.get()
        sheet.cell(row=current_row + 1, column=4).value = course_field.get()
        sheet.cell(row=current_row + 1, column=5).value = contact_no_field.get()
        sheet.cell(row=current_row + 1, column=6).value = email_id_field.get()
        sheet.cell(row=current_row + 1, column=7).value = address_field.get()
        sheet.cell(row=current_row + 1, column=8).value = strengths_field.get()
        sheet.cell(row=current_row + 1, column=9).value = weekness_field.get()

        # save the file
        wb.save('F:/COLLAGE FILES/2nd YEAR/SEM 2/Hackathon/students.xlsx')

        # set focus on the name_field box
        sname_field.focus_set()

        # call the clear() function
        clear()


# Driver code
if __name__ == "__main__":
    # create a GUI window
    root = Tk()

    # set the background colour of GUI window
    root.configure(background='light green')

    # set the title of GUI window
    root.title("registration form")

    # set the configuration of GUI window
    root.geometry("5000x3000")

    excel()

    # create a Form label
    heading = Label(root, text="Form", bg="light green",font=("Helvetica", 23))

    # create a Name label
    sname = Label(root, text="SurName", bg="light green",font=("Helvetica", 16))

    fname = Label(root, text="FirstName", bg="light green",font=("Helvetica", 16))

    # create a Course label
    lname = Label(root, text="LastName", bg="light green",font=("Helvetica", 16))

    # create a Semester label
    course = Label(root, text="Course", bg="light green",font=("Helvetica", 16))

    # create a Form No. lable

    # create a Contact No. label
    contact_no = Label(root, text="Contact No.", bg="light green",font=("Helvetica", 16))

    # create a Email id label
    email_id = Label(root, text="Email id", bg="light green",font=("Helvetica", 16))

    # create a address label
    address = Label(root, text="Address", bg="light green",font=("Helvetica", 16))

    strengths = Label(root, text="Strengths", bg="light green",font=("Helvetica", 16))

    weekness = Label(root, text="weekness", bg="light green",font=("Helvetica", 16))

    # grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    heading.grid(row=0, column=1)
    sname.grid(row=1, column=0)
    fname.grid(row=2, column=0)
    lname.grid(row=3, column=0)
    course.grid(row=4, column=0)
    contact_no.grid(row=5, column=0)
    email_id.grid(row=6, column=0)
    address.grid(row=7, column=0)
    strengths.grid(row=8, column=0)
    weekness.grid(row=9, column=0)

    # create a text entry box
    # for typing the information
    sname_field = Entry(root)
    fname_field = Entry(root)
    lname_field = Entry(root)
    course_field = Entry(root)
    contact_no_field = Entry(root)
    email_id_field = Entry(root)
    address_field = Entry(root)
    strengths_field = Entry(root)
    weekness_field = Entry(root)


    # bind method of widget is used for
    # the binding the function with the events

    # whenever the enter key is pressed
    # then call the focus1 function
    sname_field.bind("<Return>", focus1)
    fname_field.bind("<Return>", focus2)
    lname_field.bind("<Return>", focus3)
    # whenever the enter key is pressed
    # then call the focus2 function
    course_field.bind("<Return>", focus4)

    # whenever the enter key is pressed
    # then call the focus3 function

    # whenever the enter key is pressed
    # then call the focus5 function
    contact_no_field.bind("<Return>", focus5)

    # whenever the enter key is pressed
    # then call the focus6 function
    email_id_field.bind("<Return>", focus6)

    strengths_field.bind("<Return>", focus7)

    weekness_field.bind("<Return>", focus8)


# grid method is used for placing
    # the widgets at respective positions
    # in table like structure .
    sname_field.grid(row=1, column=1, ipadx="100")
    fname_field.grid(row=2, column=1, ipadx="100")
    lname_field.grid(row=3, column=1, ipadx="100")
    course_field.grid(row=4, column=1, ipadx="100")
    contact_no_field.grid(row=5, column=1, ipadx="100")
    email_id_field.grid(row=6, column=1, ipadx="100")
    address_field.grid(row=7, column=1, ipadx="100")
    strengths_field.grid(row=8, column=1, ipadx="100")
    weekness_field.grid(row=9, column=1, ipadx="100")

    # call excel function
    excel()

    # create a Submit Button and place into the root window
    submit = Button(root, text="Submit", fg="Black",
                    bg="Red", command=insert)
    submit.grid(row=10, column=1)

    # start the GUI
    root.mainloop()



def selectionES():
    if radio.get() == "Nill":
        root = Tk()
        text = Text(root)
        text.insert(INSERT, "So, you are nill at English Speaking!\nShort Talks\nShow and Tell\nRunning Dictation\nSurveys and Interviews\nFun Speaking Games\nDescriptive drawing activity\nDesert island activity\nStorytelling activity\nTrue/false storytelling\nUse English Dicstionary translate from your language")
        text.insert(END, "Bye Bye.....")
        text.pack()
        root.mainloop()
    if radio.get() == "Poor":
        root = Tk()
        text = Text(root)
        text.insert(INSERT, "So, you are poor at English Speaking!\nSet Some Speaking and Listening Goals\nListen With Your Whole Body\nPlay Listening Games\nMix Visuals With Listening\nDrawing on Demand\nWrite a Speech\n")
        text.insert(END, "Bye Bye.....")
        text.pack()
        root.mainloop()
    if radio.get() == "Average":
        root = Tk()
        text = Text(root)
        text.insert(INSERT, "So, you are average at English Speaking!\nSpeaking in Rhyme\nDictation Activities\nhttps://id.pinterest.com/ikayadin/improve-english-speaking/\nprefer above link for few games like activities\nConversation Starters for Adults\nCocktail Party to Practice Small Talk\nPartner Conversation Starters\n")
        text.insert(END, "Bye Bye.....")
        text.pack()
        root.mainloop()
    if radio.get() == "Good":
        root = Tk()
        text = Text(root)
        text.insert(INSERT, "So, you are good at English Speaking!\nClass Debate\nFilm a News Show or Skit\nMurder Mystery Party\nRecord an Interview\n")
        text.insert(END, "Bye Bye.....")
        text.pack()
        root.mainloop()
    if radio.get() == "Excellent":
        root = Tk()
        text = Text(root)
        text.insert(INSERT, "So, you are excellent at English Speaking!\nAttend interviews\nImprov Games\nWatching English movies\nAttempting Quizes\n")
        text.insert(END, "Bye Bye.....")
        text.pack()
        root.mainloop()

def selectionC():
    if radio.get() == "Nill":
        root = Tk()
        text = Text(root)
        text.insert(INSERT, "So, you are nill at Creativity!\nIncomplete figure test\n30 circles\nPaper clip test\nMusical ideas\nRe-purposed product\nDictionary story\nCompound collaboration\nBuilding blocks\nWrite poetry\n")
        text.insert(END, "Bye Bye.....")
        text.pack()
        root.mainloop()
    if radio.get() == "Poor":
        root = Tk()
        text = Text(root)
        text.insert(INSERT, "So, you are poor at Creativity!\nDraw it again\nField trip\nRead\nFree write\nStoryboard\nSCAMPER\nSix thinking hats\nQuestion assumptions\nNew out of two")
        text.insert(END, "Bye Bye.....")
        text.pack()
        root.mainloop()
    if radio.get() == "Average":
        root = Tk()
        text = Text(root)
        text.insert(INSERT, "So, you are average at Creativity!\nSet up learning activities that allow students to explore their creativity in relevant, interesting, and worthwhile ways\nValue creativity and celebrate and reward it\nTeach students the other skills they need to be creative\nRemove constraints for creativity and give the students space and a framework in which they can be creative\n")
        text.insert(END, "Bye Bye.....")
        text.pack()
        root.mainloop()
    if radio.get() == "Good":
        root = Tk()
        text = Text(root)
        text.insert(INSERT, "So, you are good at Creativity!\nhttps://www.canva.com/learn/19-ideas-to-promote-more-creativity-in-your-classroom/\nPrefer above link\n")
        text.insert(END, "Bye Bye.....")
        text.pack()
        root.mainloop()
    if radio.get() == "Excellent":
        root = Tk()
        text = Text(root)
        text.insert(INSERT, "So, you are excellent at Creativity!\nPlan projects that depend on creativity\nAlso share Ideas\n")
        text.insert(END, "Bye Bye.....")
        text.pack()
        root.mainloop()


def selectionPS():
    if radio.get() == "Nill":
        root = Tk()
        text = Text(root)
        text.insert(INSERT, "So, you are nill at Public speaking!\nhttps://www.skillsconverged.com/FreeTrainingMaterials/tabid/258/articleType/CategoryView/categoryId/98/Public-Speaking.aspx\nrefer above link\n")
        text.insert(END, "Bye Bye.....")
        text.pack()
        root.mainloop()
    if radio.get() == "Poor":
        root = Tk()
        text = Text(root)
        text.insert(INSERT, "So, you are poor at Public speaking!\nhttps://www.classcentral.com/course/public-speaking-889\n")
        text.insert(END, "Bye Bye.....")
        text.pack()
        root.mainloop()
    if radio.get() == "Average":
        root = Tk()
        text = Text(root)
        text.insert(INSERT, "So, you are average at Public speaking!\nStart a 30-day speaking challenge\nPresent at a ‘Lunch And Learn’\nPresent at local Meetups\nPresent during team meetings")
        text.insert(END, "Bye Bye.....")
        text.pack()
        root.mainloop()
    if radio.get() == "Good":
        root = Tk()
        text = Text(root)
        text.insert(INSERT, "So, you are good at Public speaking!\nhttps://www.write-out-loud.com/public-speaking-games.html\nRefer above link")
        text.insert(END, "Bye Bye.....")
        text.pack()
        root.mainloop()
    if radio.get() == "Excellent":
        root = Tk()
        text = Text(root)
        text.insert(INSERT, "So, you are excellent at Public speaking!\nhttps://medium.com/@rizwanjavaid/6-effective-ways-to-boost-your-public-speaking-skills-83ed00201d46\nRefer above link")
        text.insert(END, "Bye Bye.....")
        text.pack()
        root.mainloop()


def selectionTI():
    if radio.get() == "Nill":
        root = Tk()
        text = Text(root)
        text.insert(INSERT, "So, you are nill at Taking Intiative!\nhttps://www.mindtools.com/pages/article/initiative.html\nRefer above Link\n")
        text.insert(END, "Bye Bye.....")
        text.pack()
        root.mainloop()
    if radio.get() == "Poor":
        root = Tk()
        text = Text(root)
        text.insert(INSERT, "So, you are poor at Taking Intiative!\nhttps://www.thriveyard.com/17-tips-on-how-to-take-initiative-at-work/\nRefer above link\n")
        text.insert(END, "Bye Bye.....")
        text.pack()
        root.mainloop()
    if radio.get() == "Average":
        root = Tk()
        text = Text(root)
        text.insert(INSERT, "So, you are average at Taking Intiative!\nhttps://www.indeed.com/career-advice/career-development/ways-to-take-initiative-at-work\nRefer above link\n")
        text.insert(END, "Bye Bye.....")
        text.pack()
        root.mainloop()
    if radio.get() == "Good":
        root = Tk()
        text = Text(root)
        text.insert(INSERT, "So, you are good at Taking Intiative!\nhttps://in.pinterest.com/snyders/initiative-games/\nRefer above link\n")
        text.insert(END, "Bye Bye.....")
        text.pack()
        root.mainloop()
    if radio.get() == "Excellent":
        root = Tk()
        text = Text(root)
        text.insert(INSERT, "So, you are excellent at Taking Intiative!\nhttps://hatrabbits.com/en/10-ways-to-get-employees-to-take-initiative/\nRefer above link\n")
        text.insert(END, "Bye Bye.....")
        text.pack()
        root.mainloop()


def selectionF():
    if radio.get() == "Nill":
        root = Tk()
        text = Text(root)
        text.insert(INSERT, "So, you are nill at Focused!\nhttps://www.gettingsmart.com/2016/10/4-concentration-activities-students/\nRefer above link\n")
        text.insert(END, "Bye Bye.....")
        text.pack()
        root.mainloop()
    if radio.get() == "Poor":
        root = Tk()
        text = Text(root)
        text.insert(INSERT, "So, you are poor at Focused!\nhttps://www.edutopia.org/blog/helping-students-develop-focusing-skills-maurice-elias\n\Refer above Link\n")
        text.insert(END, "Bye Bye.....")
        text.pack()
        root.mainloop()
    if radio.get() == "Average":
        root = Tk()
        text = Text(root)
        text.insert(INSERT, "So, you are average at Focused!\nhttps://www.facultyfocus.com/articles/blended-flipped-learning/three-focusing-activities-engage-students-first-five-minutes-class/\nRefer above link\n")
        text.insert(END, "Bye Bye.....")
        text.pack()
        root.mainloop()
    if radio.get() == "Good":
        root = Tk()
        text = Text(root)
        text.insert(INSERT, "So, you are good at Focused!\nhttps://wabisabilearning.com/blogs/literacy-numeracy/future-focused-learning-activities\nRefer above link\n")
        text.insert(END, "Bye Bye.....")
        text.pack()
        root.mainloop()
    if radio.get() == "Excellent":
        root = Tk()
        text = Text(root)
        text.insert(INSERT, "So, you are excellent at Focused!\nhttps://www.hindustantimes.com/india-news/our-focus-is-on-requirement-based-skill-training-mahendra-nath-pandey/story-Ea6qTrXGP8IBcTzT4NLA5J.html\nLearn through the above link\n")
        text.insert(END, "Bye Bye.....")
        text.pack()
        root.mainloop()




top = Tk()
top.geometry("5000x3000")
lbl2 = Label(text="note: click on your choice", )
radio = StringVar()
lbl = Label(text="English Speaking:",font=("Arial Bold", 25))
top.title("English Speaking")
lbl.pack()
R1 = Radiobutton(top, text="nill",font=("Helvetica", 16), variable=radio, value="Nill", command=selectionES)
R1.pack(anchor=W)
R2 = Radiobutton(top, text="poor",font=("Helvetica", 16), variable=radio, value="Poor",command=selectionES)
R2.pack(anchor=W)
R3 = Radiobutton(top, text="average",font=("Helvetica", 16), variable=radio, value="Average",command=selectionES)
R3.pack(anchor=W)
R4 = Radiobutton(top, text="good",font=("Helvetica", 16), variable=radio, value="Good",command=selectionES)
R4.pack(anchor=W)
R5 = Radiobutton(top, text="Excellent",font=("Helvetica", 16), variable=radio, value="Excellent",command=selectionES)
R5.pack(anchor=W)
label = Label(top)
label.pack()
top.mainloop()


top = Tk()
top.geometry("5000x3000")
lbl2 = Label(text="note: click on your choice")
top.title("Creativity")
radio = StringVar()
lbl = Label(text="Creativity:",font=("Arial Bold", 25))
lbl.pack()
R1 = Radiobutton(top, text="nill",font=("Helvetica", 16), variable=radio, value="Nill",command=selectionC)
R1.pack(anchor=W)
R2 = Radiobutton(top, text="poor",font=("Helvetica", 16), variable=radio, value="Poor",command=selectionC)
R2.pack(anchor=W)
R3 = Radiobutton(top, text="average",font=("Helvetica", 16), variable=radio, value="Average",command=selectionC)
R3.pack(anchor=W)
R4 = Radiobutton(top, text="good",font=("Helvetica", 16), variable=radio, value="Good",command=selectionC)
R4.pack(anchor=W)
R5 = Radiobutton(top, text="Excellent",font=("Helvetica", 16), variable=radio, value="Excellent",command=selectionC)
R5.pack(anchor=W)
label = Label(top)
label.pack()
top.mainloop()


top = Tk()
top.geometry("5000x3000")
top.title("Public Speaking")
lbl2 = Label(text="note: click on your choice")
radio = StringVar()
lbl = Label(text="Public Speaking:",font=("Arial Bold", 25))
lbl.pack()
R1 = Radiobutton(top, text="nill",font=("Helvetica", 16), variable=radio, value="Nill",command=selectionPS)
R1.pack(anchor=W)
R2 = Radiobutton(top, text="poor",font=("Helvetica", 16), variable=radio, value="Poor",command=selectionPS)
R2.pack(anchor=W)
R3 = Radiobutton(top, text="average",font=("Helvetica", 16), variable=radio, value="Average",command=selectionPS)
R3.pack(anchor=W)
R4 = Radiobutton(top, text="good",font=("Helvetica", 16), variable=radio, value="Good",command=selectionPS)
R4.pack(anchor=W)
R5 = Radiobutton(top, text="Excellent",font=("Helvetica", 16), variable=radio, value="Excellent",command=selectionPS)
R5.pack(anchor=W)
label = Label(top)
label.pack()
top.mainloop()

top = Tk()
top.geometry("5000x3000")
top.title("Taking Intiative")
lbl2 = Label(text="note: click on your choice")
radio = StringVar()
lbl = Label(text="Taking Intiative:",font=("Arial Bold", 25))
lbl.pack()
R1 = Radiobutton(top, text="nill",font=("Helvetica", 16), variable=radio, value="Nill",command=selectionTI)
R1.pack(anchor=W)
R2 = Radiobutton(top, text="poor",font=("Helvetica", 16), variable=radio, value="Poor",command=selectionTI)
R2.pack(anchor=W)
R3 = Radiobutton(top, text="average",font=("Helvetica", 16), variable=radio, value="Average",command=selectionTI)
R3.pack(anchor=W)
R4 = Radiobutton(top, text="good",font=("Helvetica", 16), variable=radio, value="Good",command=selectionTI)
R4.pack(anchor=W)
R5 = Radiobutton(top, text="Excellent",font=("Helvetica", 16), variable=radio, value="Excellent",command=selectionTI)
R5.pack(anchor=W)
label = Label(top)
label.pack()
top.mainloop()

top = Tk()
top.geometry("5000x3000")
top.title("Focused")
lbl2 = Label(text="note: click on your choice")
radio = StringVar()
lbl = Label(text="Focused:",font=("Arial Bold", 25))
lbl.pack()
R1 = Radiobutton(top, text="nill",font=("Helvetica", 16), variable=radio, value="Nill",command=selectionF)
R1.pack(anchor=W)
R2 = Radiobutton(top, text="poor",font=("Helvetica", 16), variable=radio, value="Poor",command=selectionF)
R2.pack(anchor=W)
R3 = Radiobutton(top, text="average",font=("Helvetica", 16), variable=radio, value="Average",command=selectionF)
R3.pack(anchor=W)
R4 = Radiobutton(top, text="good",font=("Helvetica", 16), variable=radio, value="Good",command=selectionF)
R4.pack(anchor=W)
R5 = Radiobutton(top, text="Excellent",font=("Helvetica", 16), variable=radio, value="Excellent",command=selectionF)
R5.pack(anchor=W)
label = Label(top)
label.pack()
top.mainloop()